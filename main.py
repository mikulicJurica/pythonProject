import openpyxl
from openpyxl.styles import Alignment, PatternFill, Border, Side
from gurobipy import *
from decimal import Decimal
import matplotlib.pyplot as plt

# Variables for storing data from provided excel file
hour = []  # hour array
PV_production = []  # solar panels production array; production in certain hour
Load = []  # consumption array; consumption in certain hour
Price = []  # price of energy in given hour; euro per kWh

# (openpyxl configuration) - Importing excel data
Input_data_file = openpyxl.load_workbook("Podaci_9.xlsx")
Input_data_sheet = Input_data_file.active

# Reading data from excel and assigning data to correct variables
for col in Input_data_sheet.iter_rows(min_row=2, max_col=8, max_row=169, values_only=True):
    hour.append(col[0])
    PV_production.append(col[1])
    Load.append(round(col[4], 1))
    Price.append(col[7])

Battery_capacity_optimal = 0  # variable for calculating battery storage capacity

# Variables - WITHOUT a battery
Nett_ene_from_network_NO_Battery = []  # total nett energy with no battery
Cost_in_hour_NO_Battery = []  # cost in certain hour with no battery
Total_cost_NO_Battery = 0  # variable for total cost with no battery storage

# Variables - WITH a battery
P_ch_amount = []  # battery charging amount in given hour
P_dis_amount = []  # battery discharging amount in given hour
SoC_amount = []  # battery SoC amount in given hour
Nett_ene_from_network_WITH_Battery = []  # total nett energy with battery
Cost_in_hour_WITH_Battery = []  # cost in certain hour with battery
Total_cost_WITH_Battery = 0  # variable for total cost with battery storage

# Making calculations for system WITHOUT a battery and calculation of optimal battery capacity
for i in range(len(hour)):
    # WITHOUT a battery
    tmp_nett_ene = Load[i] - PV_production[i]
    Nett_ene_from_network_NO_Battery.append(tmp_nett_ene)
    Cost_in_hour_NO_Battery.append(Price[i] * Nett_ene_from_network_NO_Battery[i])
    Total_cost_NO_Battery += Cost_in_hour_NO_Battery[i]

    # Calculate optimal battery capacity for worst case scenario (detecting the maximum power gap between the load and the PV-s)
    if tmp_nett_ene > Battery_capacity_optimal:
        Battery_capacity_optimal = tmp_nett_ene

# Battery storage definition - uncomment for desired data
Battery_capacity = Battery_capacity_optimal
#Battery_capacity = 20
Battery_price_per_kWh = 200
#Battery_capacity = 13.5  # Tesla Powerwall
#Battery_price_per_kWh = 778  # Tesla Powerwall

# Optimisation part - Battery storage system
m = Model("Battery_storage")  # creating new model

# Battery model - specific variables
ni_ch = 1  # charging coefficient (not defined in this task)
ni_dis = 1  # discharging coefficient (not defined in this task)

# Battery degradation
Battery_degradation_percentage = 0.02  # battery degradation in % after each hour
Battery_degradation_amount = (Battery_degradation_percentage / 100) * Battery_capacity  # amount of degradation after each hour
Battery_degradation = []

for i in range(len(hour)):
    Battery_degradation.append(round((i * Battery_degradation_amount), 4))

# Variables (gurobipy)
P_ch = {}  # charging value
P_dis = {}  # discharging value
x_ch = {}  # charging indicator (logical: 0 or 1)
x_dis = {}  # discharging indicator (logical: 0 or 1)
SoC = {}  # state of charge

# Create variables (gurobipy)
for i in range(len(hour)):
    P_ch[i] = m.addVar(name='P_ch')
    P_dis[i] = m.addVar(name='P_dis')
    x_ch[i] = m.addVar(vtype=GRB.BINARY, name='x_ch')
    x_dis[i] = m.addVar(vtype=GRB.BINARY, name='x_dis')
    SoC[i] = m.addVar(name='SoC')

# Defining Constrains in certain hours
m.addConstr(SoC[0] == 0)  # battery is empty at the beginning
m.addConstr(SoC[167] == 0)  # battery must be empty at the end
m.addConstr(P_dis[0] == 0)  # discharging is not possible at the beginning

# Defining Constrains in whole time period
for j in range(len(hour) - 1):
    i = j + 1  # skip first case (already defined)
    m.addConstr(x_ch[i] + x_dis[i] <= 1)  # not possible to charge and discharge battery at same time
    m.addConstr(P_ch[i] <= (Battery_capacity - Battery_degradation[i]) * x_ch[i])  # charging limitation
    m.addConstr(P_dis[i] <= SoC[i - 1] * x_dis[i])  # discharging limitation
    m.addConstr(P_dis[i] <= Load[i] - PV_production[i])  # not possible to discharge more than is actually needed
    m.addConstr(SoC[i] <= Battery_capacity - Battery_degradation[i])  # preventing overcharge

    m.addConstr(SoC[i] == SoC[i - 1] + (ni_ch * P_ch[i]) - (P_dis[i] / ni_dis))  # battery SoC model

# Function for cost minimization
def goal_function():
    total_cost_with_battery = 0  # variable for total cost with battery storage
    for i in range(len(hour)):
        total_cost_with_battery = total_cost_with_battery + (Load[i] - PV_production[i] + P_ch[i] - P_dis[i]) * Price[i]
    return total_cost_with_battery


# Start optimisation (gurobipy)
m.setObjective(goal_function(), GRB.MINIMIZE)
m.optimize()

# Store optimised data to appropriate variables
for v in m.getVars():
    if v.varname == 'P_ch':
        P_ch_amount.append(v.x)
    elif v.varname == 'P_dis':
        if v.x <= Battery_degradation_percentage:
            P_dis_amount.append(0)
        else:
            P_dis_amount.append(v.x)
    elif v.varname == 'SoC':
        SoC_amount.append(v.x)

# Making calculations for system WITH a battery - for nett ene. consumed from the network and for cost per hour
for i in range(len(hour)):
    Nett_ene_from_network_WITH_Battery.append(round(Load[i] - PV_production[i] + P_ch_amount[i] - P_dis_amount[i], 2))
    Cost_in_hour_WITH_Battery.append(round(Nett_ene_from_network_WITH_Battery[i] * Price[i], 2))
    Total_cost_WITH_Battery += Cost_in_hour_WITH_Battery[i]

# (openpyxl configuration) - Exporting data to excel file
Export_data_file = openpyxl.Workbook()
Export_data_sheet = Export_data_file.active

# Cell merging
Export_data_sheet.merge_cells('A1:D1')
Export_data_sheet.merge_cells('E1:F1')
Export_data_sheet.merge_cells('G1:K1')

Export_data_sheet.merge_cells('A172:C172')
Export_data_sheet.merge_cells('D172:E172')
Export_data_sheet.merge_cells('A173:C173')
Export_data_sheet.merge_cells('D173:E173')
Export_data_sheet.merge_cells('A174:C174')
Export_data_sheet.merge_cells('D174:E174')

# Cell naming
Export_data_sheet['A1'].value = "Input data"
Export_data_sheet['E1'].value = "WITHOUT use of a battery"
Export_data_sheet['G1'].value = "WITH use of a battery"

Export_data_sheet['A2'].value = "Hour"
Export_data_sheet['B2'].value = "PV production [kWh]"
Export_data_sheet['C2'].value = "Load [kW]"
Export_data_sheet['D2'].value = "Price [Eur/kWh]"
Export_data_sheet['E2'].value = "Nett ene. consumed from grid [kW]"
Export_data_sheet['F2'].value = "Cost in given hour [Eur]"

Export_data_sheet['G2'].value = "P charging [kW]"
Export_data_sheet['H2'].value = "P discharging [kW]"
Export_data_sheet['I2'].value = "SoC (Max. Soc) [kWh]"
Export_data_sheet['J2'].value = "Nett ene. consumed from grid [kW]"
Export_data_sheet['K2'].value = "Cost in given hour [Eur]"

Export_data_sheet['A172'].value = "Battery capacity [kWh]:"
Export_data_sheet['A173'].value = "Total energy cost (NO battery storage) [Eur]:"
Export_data_sheet['A174'].value = "Total energy cost (WITH battery storage) [Eur]:"

# Table styling - calculated data
Export_data_sheet['A172'].alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
Export_data_sheet['A172'].fill = PatternFill(start_color='00ABEBC6', end_color='00ABEBC6', fill_type='solid')
Export_data_sheet['D172'].alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
Export_data_sheet['D172'].fill = PatternFill(start_color='00ABEBC6', end_color='00ABEBC6', fill_type='solid')

Export_data_sheet['A173'].alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
Export_data_sheet['A173'].fill = PatternFill(start_color='00F9E79F', end_color='00F9E79F', fill_type='solid')
Export_data_sheet['D173'].alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
Export_data_sheet['D173'].fill = PatternFill(start_color='00F9E79F', end_color='00F9E79F', fill_type='solid')

Export_data_sheet['A174'].alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
Export_data_sheet['A174'].fill = PatternFill(start_color='0085C1E9', end_color='0085C1E9', fill_type='solid')
Export_data_sheet['D174'].alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
Export_data_sheet['D174'].fill = PatternFill(start_color='0085C1E9', end_color='0085C1E9', fill_type='solid')

# Table styling - rest of table
for i in range(11):
    for j in range(170):
        Export_data_sheet.cell(row=j + 1, column=i + 1).alignment = Alignment(vertical="center", horizontal="center", wrapText=True)
        Export_data_sheet.cell(row=j + 1, column=i + 1).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        if 0 <= i <= 3:
            Export_data_sheet.cell(row=j + 1, column=i + 1).fill = PatternFill(start_color='00ABEBC6', end_color='00ABEBC6', fill_type='solid')
        elif 4 <= i <= 5:
            Export_data_sheet.cell(row=j + 1, column=i + 1).fill = PatternFill(start_color='00F9E79F', end_color='00F9E79F', fill_type='solid')
        elif 6 <= i <= 10:
            Export_data_sheet.cell(row=j + 1, column=i + 1).fill = PatternFill(start_color='0085C1E9', end_color='0085C1E9', fill_type='solid')

# For loop for storing all the data into new excel file
for i in range(len(hour)):
    # Input data
    Export_data_sheet.cell(row=i + 3, column=1).value = hour[i]
    Export_data_sheet.cell(row=i + 3, column=2).value = PV_production[i]
    Export_data_sheet.cell(row=i + 3, column=3).value = Load[i]
    Export_data_sheet.cell(row=i + 3, column=4).value = Price[i]

    # Data WITHOUT a battery storage
    Export_data_sheet.cell(row=i + 3, column=5).value = Nett_ene_from_network_NO_Battery[i]
    Export_data_sheet.cell(row=i + 3, column=6).value = Cost_in_hour_NO_Battery[i]

    # Data WITH a battery storage
    Export_data_sheet.cell(row=i + 3, column=7).value = Decimal('%.3f' % (P_ch_amount[i])).normalize()
    Export_data_sheet.cell(row=i + 3, column=8).value = Decimal('%.3f' % (P_dis_amount[i])).normalize()

    SoC_normalize = str(Decimal('%.3f' % (SoC_amount[i])).normalize())
    SoC_max_normalize = str(Decimal('%.3f' % (Battery_capacity - Battery_degradation[i])).normalize())
    Export_data_sheet.cell(row=i + 3, column=9).value = ("%s\n(%s)" % (SoC_normalize, SoC_max_normalize))

    Export_data_sheet.cell(row=i + 3, column=10).value = Nett_ene_from_network_WITH_Battery[i]
    Export_data_sheet.cell(row=i + 3, column=11).value = Cost_in_hour_WITH_Battery[i]

# Storing calculated data into excel file
Export_data_sheet['D172'].value = Battery_capacity
Export_data_sheet['D173'].value = round(Total_cost_NO_Battery, 2)
Export_data_sheet['D174'].value = round(Total_cost_WITH_Battery, 2)

Export_data_file.save(filename='Microgrid_report.xlsx')

# Graphs
plt.style.use('seaborn-darkgrid')

# Graph 1 - Battery degradation
tmp_capacity = []

for i in range(len(hour)):
    tmp_capacity.append(round(Battery_capacity - Battery_degradation[i], 4))

fig, ax = plt.subplots()
fig.set_size_inches(12, 6)
ax.plot(hour, tmp_capacity, label='State of charge')
ax.axis([0, 171, Battery_capacity - Battery_degradation[len(hour) - 1] - 4, Battery_capacity + 1])
ax.set_xlabel('Time [h]', fontsize=15)
ax.set_ylabel('Max. capacity [kWh]', fontsize=15)
ax.legend(fontsize=13)
ax.set_title("Battery degradation", fontsize=15, fontweight='bold')
fig.savefig('1_Battery_degradation.png', dpi=100)

# Graph 2 - Nett energy consumed by microgrid per hour from distribution network
fig, ax = plt.subplots()
fig.set_size_inches(12, 6)
ax.plot(hour, Nett_ene_from_network_WITH_Battery, label='WITH BATTERY STORAGE')
ax.plot(hour, Nett_ene_from_network_NO_Battery, label='NO BATTERY STORAGE')
ax.axis([0, 171, 0, round(max(Nett_ene_from_network_WITH_Battery)) + 5])
ax.set_xlabel('Time [h]', fontsize=15)
ax.set_ylabel('Energy [kWh]', fontsize=15)
ax.legend(fontsize=13)
ax.set_title("Nett energy consumed by microgrid per hour from distribution network", fontsize=15, fontweight='bold')
fig.savefig('2_Nett_energy_consumption.png', dpi=100)

# Graph 3 - Overview of network prices in the hours when the battery is being charged / discharged
fig, ax = plt.subplots()
ax.plot(hour, Price)
fig.set_size_inches(12, 6)

# Scatter variables
h = []
k = []
r = []
s = []

for i in range(len(hour)):
    if P_ch_amount[i] > 0:
        h.append(hour[i])
        k.append(P_ch_amount[i] / P_ch_amount[i] * Price[i])
    if P_dis_amount[i] > 0:
        r.append(hour[i])
        s.append(P_dis_amount[i] / P_dis_amount[i] * Price[i])

ax.scatter(h, k, s=50, facecolor='g', edgecolor='k', label='Battery charging')
ax.scatter(r, s, marker='s', facecolor='r', edgecolor='k', label='Battery discharging')
ax.legend(fontsize=13)
ax.axis([0, 171, 0, round(max(Price)) + 20])
ax.set_xlabel('Time [h]', fontsize=15)
ax.set_ylabel('Price per hour [Eur]', fontsize=15)
ax.set_title("Spot indicators of charging/discharging of a battery on price graph", fontsize=15, fontweight='bold')
fig.savefig('3_ch_dis_indicators.png', dpi=100)

# Graph 4 - Comparison of cumulative cost during microgrid operation with and without a battery storage
Cumulative_cost_NO_Battery = []
Cumulative_cost_WITH_Battery = []
tmp_NO_Battery = 0
tmp_WITH_Battery = 0

for i in range(len(hour)):
    tmp_NO_Battery = tmp_NO_Battery + Cost_in_hour_NO_Battery[i]
    tmp_WITH_Battery = tmp_WITH_Battery + Cost_in_hour_WITH_Battery[i]
    Cumulative_cost_NO_Battery.append(round(tmp_NO_Battery, 2))
    Cumulative_cost_WITH_Battery.append(round(tmp_WITH_Battery, 2))

fig, ax = plt.subplots()
fig.set_size_inches(12, 6)
ax.plot(hour, Cumulative_cost_WITH_Battery, label='WITH BATTERY STORAGE')
ax.plot(hour, Cumulative_cost_NO_Battery, label='NO BATTERY STORAGE')
ax.legend(fontsize=13)
ax.axis([0, 170, 0, round(max(Cumulative_cost_NO_Battery)) + 1000])
ax.set_xlabel('Time [h]', fontsize=13)
ax.set_ylabel('Cost [Eur]', fontsize=13)
ax.set_title("Comparison of cumulative cost during microgrid operation with and without a battery storage", fontsize=15, fontweight='bold')
fig.savefig('4_Cumulative_cost.png', dpi=100)

# Bar plot - Comparison of total costs
fig = plt.figure(figsize=(12, 6))
names = ['WITHOUT Battery', 'WITH Battery', 'WITH Battery (Battery investment included)']
values = [Total_cost_NO_Battery, Total_cost_WITH_Battery, Total_cost_WITH_Battery + (Battery_capacity * Battery_price_per_kWh)]
c = ['red', 'blue', 'yellow']
plt.bar_label(plt.bar(names, height=values, color=c))
plt.ylabel('Cost [Eur]', fontsize=13)
plt.title("Total microgrid cost in given timeframe", fontsize=15, fontweight='bold')
plt.savefig('5_bar_cost.png', dpi=100)
